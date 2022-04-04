# -*- coding: utf-8 -*-
import numpy as np
import matplotlib.pyplot as plt
import time
import pandas as pd
import openpyxl
from algorithm import bag
from algorithm import show
from algorithm import QuickSort
from algorithm import KnapSack
from algorithm import test
from sql import read


# excel文件输入
def out_excel(x):
    f = open('out.txt', 'w', encoding='utf-8')
    for i in x:
        f.write(str(i) + str(x[i])+ '\n')
    f.close()
    data = pd.read_csv("out.txt", sep="\t")
    wb = openpyxl.load_workbook(r'out.xlsx')
    ws = wb['Sheet1']
    # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
    ws.cell(row=1, column=2).value ='物品'
    ws.cell(row=1, column=3).value ='是否装入'
    for i in range(1, len(x) + 1):
        distance = x[i - 1]
        # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
        ws.cell(row=i + 1, column=3).value = distance
        distance = i
        # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
        ws.cell(row=i + 1, column=2).value = distance
    # 保存操作
    wb.save(r'out1.xlsx')



if __name__ == "__main__":
    file = ' '
    option = int(input('数据集(0-9):'))
    if option == 0:
        file = 'data\\beibao0.in'
    if option == 1:
        file = 'data\\beibao1.in'
    if option == 2:
        file = 'data\\beibao2.in'
    if option == 3:
        file = 'data\\beibao3.in'
    if option == 4:
        file = 'data\\beibao4.in'
    if option == 5:
        file = 'data\\beibao5.in'
    if option == 6:
        file = 'data\\beibao6.in'
    if option == 7:
        file = 'data\\beibao7.in'
    if option == 8:
        file = 'data\\beibao8.in'
    if option == 9:
        file = 'data\\beibao9.in'
    # print(file)
    read(file)  # sql读beibao文件
    f = open(file, 'r', encoding = 'utf-8')
    arr = []
    for m1 in f:
        m2 = m1.strip("\n")
        arr.append(m2)
    del f, m1
    arr_weight = []
    arr_value = []
    for row in arr:
        arr_weight.append(row.split()[0])
        arr_value.append(row.split()[1])  # 用空格将列表项拆分开
    arr_value = list(map(int, arr_value))  # 列表项为字符串，强制转换为整数型
    arr_weight = list(map(int, arr_weight))
    nums = arr_value[0]
    capacity = arr_weight[0]

    # 删除第一行物品个数和背包容量，列表整体前移
    for i in range(len(arr_value)-1):
        arr_value[i] = arr_value[i+1]
        arr_weight[i] = arr_weight[i+1]
    del arr_value[nums-1]
    del arr_weight[nums-1]
    # print(arr_value)
    # print(arr_weight)

    # 画散点图
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    # matplotlib画图中中文显示会有问题，需要这两行设置默认字体
    # rcParams用来设置画图时的一些基本参数

    # 横轴与纵轴名称及大小
    plt.xlabel('Weight')
    plt.ylabel('Value')
    plt.xlim(xmax=110, xmin=0)
    plt.ylim(ymax=110, ymin=0)

    # 点的颜色
    colors2 = '#DC143C'

    # 点面积
    area = np.pi * 2 ** 2

    plt.scatter(arr_weight, arr_value, s=area, c=colors2, alpha=0.4, label='物品')
    plt.legend()  # 显示字符或表达式向量
    plt.savefig(r'D:\1.png', dpi=300)
    plt.show()

    # 计算性价比
    ratio = []
    for i in range(len(arr_value)):
        b = round((arr_value[i]/arr_weight[i]), 4)
        ratio.append(b)
        # print('价值：' + str(arr_value[i]) + ' ' + '重量：' + str(arr_weight[i]) + ' ' + '性价比：' + str(b))

    # 利用快速排序算法，按性价比非递增排序
    QuickSort(ratio, arr_value, arr_weight, 0, nums-1)
    for i in range(len(arr_value)):
        print('价值：' + str(arr_value[i]) + ' ' + '重量：' + str(arr_weight[i]) + ' ' + '性价比：' + str(ratio[i]))

    # 结果写入文件
    f = open('out.txt', 'w', encoding='utf-8')
    f.write('性价比')
    f.write('\n')
    for i in ratio:
        f.write(str(i) + '\n')
    f.close()
    data = pd.read_csv("out.txt", sep="\t")
    data.to_excel("out.xlsx", index=False)

    option = int(input('1贪心法 2动态规划法 3回溯法'+'\n'))
    # option = 3
    start = time.perf_counter()
    # while option:
    # 贪心法
    if option == 1:
        value = KnapSack(arr_weight, arr_value, nums, capacity)
        value = round(value, 4)
        print(value)

    # 动态规划法
    if option == 2:
        value = bag(nums, capacity, arr_weight, arr_value)
        show(nums, capacity, arr_weight, value)
    # print()

    # 回溯法
    if option == 3:
        test(capacity, arr_weight, arr_value)
        print()

    end = time.perf_counter()
    print()
    print('running time: %.8s s' % ((end - start)))

    # 结果写入文件
    f = open('out.txt', 'w', encoding='utf-8')
    f.write(str(end-start)+ 's'+ '\n')
    f.close()



